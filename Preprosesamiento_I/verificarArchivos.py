from os import walk
from pptx import Presentation
import openpyxl
import os

def leeExcell(ruta):
	doc = openpyxl.load_workbook(ruta)
	a=doc.get_sheet_names()
	#print(a[0])
	#print(type(a))
	#linea= doc.get_sheet_by_name('Sheet1')
	linea=a[0]
	return linea

def leetxt(ruta):
	doc=open(ruta)
	linea=doc.readline()
	return linea
def leePdf(ruta):
	#doc = openpyxl.load_workbook(ruta)
	#a=doc.get_sheet_names()
	#print(a)
	#linea= doc.get_sheet_by_name('Sheet1')
	linea="Entre pdf"
	return linea

def leePptx(ruta):
	#print (ruta)
	prs = Presentation(ruta)
	
	linea="Entre pptx" 
	return linea

def leePpt(ruta):
	
	#print (ruta)
	#prs = Presentation(ruta)
	
	linea="Entre ppt" 
	return linea

def leedoc(ruta):
	#doc = openpyxl.load_workbook(ruta)
	#a=doc.get_sheet_names()
	#print(a)
	#linea= doc.get_sheet_by_name('Sheet1')
	linea="Entre pdoc"
	return linea


ruta ="F:/temp\Software\prueba"


#print (ruta)
for (path,ficheros,archivos) in walk(ruta):
	#print ("Ruta")
	#print ( path)
	#print("carpetas")
	#print ( ficheros)
	#print ("archivos")
	#print ( archivos)
	for archivo in archivos:
		
		(nombre,extencion)=os.path.splitext(archivo)
		#print(extencion)
		tipos ={ '.xlsx':leeExcell,'.txt':leetxt,'.py':leetxt, '.pdf':leePdf,'.ppt':leePpt ,'.pptx':leePptx ,'.docx':leedoc   }
		try:
			linea=tipos[extencion](path + "/" + archivo)
			
		except:
			print("extencion sin validar:" + extencion)
			raise
		
		if  linea=='':
		   print(path + "/" + archivo) 
		else:
			print( linea + " " + path + "/" + archivo) 
